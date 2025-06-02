import io
import os
import re
import sys
import glob
from google.cloud import vision
import pandas as pd
from openpyxl import load_workbook

# Set up the Vision API client
client = vision.ImageAnnotatorClient()

# Function to find the most recent image file in the directory
def find_latest_image(directory):
    # Look for any supported image files in the directory
    files = glob.glob(os.path.join(directory, '*.[jp][pn]g'))  # Match .jpg, .jpeg, .png
    if not files:
        raise FileNotFoundError("No image files found in the directory.")
    latest_file = max(files, key=os.path.getctime)
    return latest_file

# Determine the image path
directory = "/Users/zishan/Desktop/Excel_Test"
image_path = sys.argv[1] if len(sys.argv) > 1 else find_latest_image(directory)
excel_path = os.path.join(directory, "data.xlsx")

def detect_text(image_path):
    """Detects text in the file."""
    with io.open(image_path, 'rb') as image_file:
        content = image_file.read()

    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations

    if response.error.message:
        raise Exception(f'{response.error.message}')

    return texts

def parse_driver_license_text(texts):
    """Parse the extracted text and map it to the relevant fields."""
    data = {
        "Phone": "",
        "Date": pd.to_datetime("today").strftime("%m/%d/%Y"),
        "Have you Been Here Before?": "",
        "First Name": "",
        "Last Name": "",
        "Date Of Birth": "",
        "Email Address": "",
        "Address": "",
        "City": "",
        "State": "",
        "Zip Code": "",
        "Type Of Service": "",
        "Payment Type": "",
        "Amount": "",
        "Comments": "",
        "Image": ""  # Placeholder for the image link
    }

    # Join all text into a single string and lowercase it
    full_text = ' '.join([text.description for text in texts])

    # Debugging output
    print("Full extracted text:", full_text)

    # Extract DOB by looking for "DOB:" or "Date of Birth:"
    dob_match = re.search(r'(DOB[:\s]*|Date of Birth[:\s]*)(\d{2}/\d{2}/\d{4})', full_text)
    if dob_match:
        data["Date Of Birth"] = dob_match.group(2)

    # Extract Last Name: Strictly using labels
    last_name_labels = ['Family Name', 'Surname', 'Last Name', 'LN', r'\b1\s+([^\s]+)']
    for label in last_name_labels:
        last_name_match = re.search(rf'{label}[:\s]*([^\n]+)', full_text, re.IGNORECASE)
        if last_name_match:
            data["Last Name"] = last_name_match.group(1).strip().capitalize()
            break

    # Extract First Name: Strictly using labels
    first_name_labels = ['Given Names', 'First Name', 'FN', r'\b2\s+([^\s]+)']
    for label in first_name_labels:
        first_name_match = re.search(rf'{label}[:\s]*([^\n]+)', full_text, re.IGNORECASE)
        if first_name_match:
            data["First Name"] = first_name_match.group(1).strip().capitalize()
            break

    # Debugging output for names
    print(f"First Name: {data['First Name']}, Last Name: {data['Last Name']}")

    # Enhanced Address Extraction with case variations
    address_keywords = [
        'Circle', 'CIR', 'Drive', 'DR', 'Boulevard', 'BLVD', 'Street', 'ST', 
        'Avenue', 'AVE', 'Road', 'RD', 'Lane', 'LN', 'Court', 'CT', 'Place', 'PL',
        'Terrace', 'TER', 'Way', 'WY'
    ]
    address_match = None

    # Search for "Address" or "8" with context-sensitive parsing to avoid accidental matches
    address_regex = r'\b(8\s+[^\n]*?\b(?:' + '|'.join(address_keywords) + r')\b[^\n]*)'
    address_match = re.search(address_regex, full_text, re.IGNORECASE)
    if not address_match:
        # Fallback to detecting "address" keyword
        address_match = re.search(r'Address[:\s]*([^\n]+)', full_text, re.IGNORECASE)

    if address_match:
        data["Address"] = address_match.group(1).strip().title()

    # Extract ZIP Code
    zip_code_match = re.search(r'\b\d{5}(?:-\d{4})?\b', full_text)  # Matches "ZIP Code"
    if zip_code_match:
        data["Zip Code"] = zip_code_match.group(0)

    return data

def update_excel_with_customer_data(data, excel_path, image_path):
    """Update the Excel sheet with the extracted data and insert the image link."""
    sheet_name = 'Sheet1'
    
    if not os.path.exists(excel_path):
        # Create a new DataFrame and Excel file if it doesn't exist
        df = pd.DataFrame(columns=data.keys())
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
    
    # Load the workbook and sheet
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name]
    
    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Insert data into the Excel sheet
    for col, key in enumerate(data.keys(), start=1):
        sheet.cell(row=next_row, column=col, value=data[key])
    
    # Insert the hyperlink to the image in the "Image" column
    image_link_cell = sheet.cell(row=next_row, column=len(data.keys()))
    image_link_cell.value = "View Image"
    image_link_cell.hyperlink = image_path
    image_link_cell.style = "Hyperlink"
    
    # Save the updated Excel file
    workbook.save(excel_path)

def process_driver_license(image_path, excel_path):
    texts = detect_text(image_path)
    data = parse_driver_license_text(texts)
    update_excel_with_customer_data(data, excel_path, image_path)

if __name__ == '__main__':
    process_driver_license(image_path, excel_path)
